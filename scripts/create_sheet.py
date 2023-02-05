import openpyxl
import re
from copy import copy
from collections import defaultdict

from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker
from openpyxl.drawing.xdr import XDRPoint2D, XDRPositiveSize2D
from openpyxl.utils.units import pixels_to_EMU, cm_to_EMU

ALPH = "ABCDEFGHIJKLMNOPQRSTUVWXYZ"
def get_col(index):
    index -= 1 # 0-based logic is easier
    if index < len(ALPH):
        return ALPH[index]
    return get_col(index // len(ALPH) + 1) + ALPH[index % len(ALPH)]

def get_index(col):
    return (ALPH.index(col[-1]) + 1) + (
        26 * get_index(col[:-1])
        if len(col) > 1
        else 0
    )

def to_tuple(key):
    """Turns D32 into (4, 32)"""
    g = re.match(r"^(?P<col>[A-Z]*)(?P<row>[0-9]*)$", key)
    return (get_index(g.group("col")), int(g.group("row")))

def create_sheets(students, sheets, conf, images):

    for student_info, sheet_info in zip(students, sheets):
        workbook = openpyxl.load_workbook(conf["workbook"])
        question_sheet = workbook[conf["question_sheet"]]
        rubric_sheet = workbook[conf["rubric_sheet"]]
        workbooks = {
            "questions": question_sheet,
            "rubric": rubric_sheet,
        }

        _draw_computed(student_info, sheet_info, conf.get("computed_cols", []), workbooks)

        _draw_questions(student_info, sheet_info, workbooks, conf, images)

        workbook.save(student_info["path"])

def _draw_computed(student_info, sheet_info, computed_cols, workbooks):
    for entry in computed_cols:
        if entry["sheet"] not in workbooks:
            raise ValueError("Wrong Sheet value.")
        sheet = workbooks[entry["sheet"]]
        sheet[entry["key"]].value = entry["expr"].format(info=student_info)

def _draw_questions(student_info, sheet_info, workbooks, conf, images):
    merges = defaultdict(set)
    for question, qloc, rloc in zip(sheet_info, conf["question_locations"], conf["rubric_locations"]):
        question_vals = list(map(lambda x: (to_tuple(x[0]), x[1]), question["question_cells"].items()))
        question_vals.sort()

        min_rc = question_vals[0][0]

        # Figure out global offset.
        question_rc = to_tuple(qloc)
        offset = (question_rc[0] - min_rc[0], question_rc[1] - min_rc[1])

        # Copy question contests
        for loc_tuple, info in question_vals:
            actual_tuple = (loc_tuple[0] + offset[0], loc_tuple[1] + offset[1])
            actual_coord = f"{get_col(actual_tuple[0])}{actual_tuple[1]}"
            new_cell = workbooks["questions"][actual_coord]
            if "style" in info:
                new_cell.font = copy(info["style"]["font"])
                new_cell.border = copy(info["style"]["border"])
                new_cell.fill = copy(info["style"]["fill"])
                new_cell.number_format = copy(info["style"]["number_format"])
                new_cell.protection = copy(info["style"]["protection"])
                new_cell.alignment = copy(info["style"]["alignment"])
            if info["merged"] is not None:
                merge_tuple = to_tuple(info["merged"])
                new_tuple = (merge_tuple[0] + offset[0], merge_tuple[1] + offset[1])
                merges[new_tuple].add(actual_tuple)
            elif "value" in info:
                if info["value"].startswith("img:"):
                    key = info["value"][4:]
                    image = openpyxl.drawing.image.Image(key)
                    details = images[key]
                    old_height = image.height
                    old_width = image.width
                    if "preffered_width" in details:
                        image.width = details["preffered_width"]
                        if not "preffered_height" in details:
                            image.height *= details["preffered_width"] / old_width
                    if "preffered_height" in details:
                        image.height = details["preffered_height"]
                        if not "preffered_width" in details:
                            image.width *= details["preffered_height"] / old_height
                    size = XDRPositiveSize2D(pixels_to_EMU(image.width), pixels_to_EMU(image.height))
                    marker = AnchorMarker(
                        col=actual_tuple[0],
                        colOff=cm_to_EMU(details.get("offset_x", 0) * image.width * 0.02666),
                        row=actual_tuple[1],
                        rowOff=cm_to_EMU(details.get("offset_y", 0) * image.height * 0.02666)
                    )
                    image.anchor = OneCellAnchor(_from=marker, ext=size)
                    workbooks["questions"].add_image(image)
                else:
                    new_cell.value = info["value"]

        # Add short description to rubric.
        shortened = question_vals[0][1]["value"]
        if len(shortened) > 33:
            shortened = f"{shortened[:30]}+..."
        workbooks["rubric"][rloc].value = shortened

    for parent, children in merges.items():
        workbooks["questions"].merge_cells(start_row=parent[1], start_column=parent[0], end_row=max(children)[1], end_column=max(children)[0])
